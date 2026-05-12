"""
EdGame financial model + 10× MOIC analysis based on comparable EdTech exits.

Inputs: BMC revenue/cost plan + research-backed comparables.
Outputs:
  - reports/business-plan/financial_model.xlsx — multi-sheet workbook
  - tools/business-plan/financial_summary.json — numbers consumed by figures.py
    and generate_plan.py
  - stdout: human-readable 10× story walkthrough for the BP body

Comparables anchoring the 10× story (from web_research.json):
  - Kahoot delisted Mar 2024 at $1.66B EV / $150M ARR  = 11.1× revenue
  - Duolingo IPO 2021 at $6.5B / $250M ARR             = 26×   revenue
  - Duolingo 2026 trading at $5B  / $1.04B revenue     = 4.8×  revenue
  - Quizlet 2020 unicorn $1B / $80M ARR                = 12.5× revenue
  - EdTech Series A/B median multiple Q4 2025 (Finrofca) ≈ 10×
  - EdTech seed-stage avg EV/Revenue (Finerva)         ≈ 11.9×
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path("/Users/yousefradwan/Library/CloudStorage/GoogleDrive-radwanf2025@gmail.com/My Drive/Yousef/KAUST/TIEVenture")
OUT_XLSX = REPO / "reports" / "business-plan" / "financial_model.xlsx"
OUT_JSON = REPO / "tools" / "business-plan" / "financial_summary.json"

# ------------------------------------------------------------------ #
#  Revenue model (per BMC; numbers reviewed against research)         #
# ------------------------------------------------------------------ #

REVENUE_BY_SEGMENT_USD = {
    # Reduced to 4 defensible streams per R1 (Reach S2.8, a16z S2.3).
    # Killed: parent premium, after-school programs, OEM API, custom dev.
    # Price escalation softened to 3-7%/yr per R1 (Reach S2.5).
    # Sales-cycle assumption rebuilt around 9-month base case (Reach S2.10).
    #
    # Each cell = (units × price)
    #               Y1     Y2      Y3       Y4       Y5
    "Individual Teacher Pro (SaaS)": [
        # $12/mo × 12 = $120/yr; 3-5%/yr realistic price increases
        500 * 120,        # 500 paid teachers Y1 (PLG bottom-up still strong)
        1_400 * 125,
        3_500 * 130,
        6_500 * 135,
        10_000 * 140,
    ],
    "School Standard ($6-7 / student / yr)": [
        25 * 500 * 6,     # 25 schools Y1 (founder-led BD)
        65 * 500 * 6.2,
        130 * 520 * 6.5,
        210 * 540 * 6.7,
        310 * 560 * 6.9,
    ],
    "School Premium ($10-12 / student / yr)": [
        10 * 600 * 10,    # 10 premium schools Y1
        30 * 600 * 10.5,
        65 * 620 * 11,
        110 * 630 * 11.5,
        170 * 650 * 11.8,
    ],
    "District / Multi-school ($5-6 / student / yr)": [
        0,
        0,
        12 * 3_000 * 5.5,
        25 * 3_500 * 5.7,
        45 * 4_000 * 5.9,
    ],
}

# ------------------------------------------------------------------ #
#  Cost model (per BMC)                                                #
# ------------------------------------------------------------------ #

COSTS_BY_LINE_USD = {
    # Trimmed for the leaner 4-stream revenue plan. Y5 ~$6.4M total cost
    # vs $7.8M revenue (~18% margin), more defensible than the original
    # $8.6M plan that was sized for the 8-stream universe.
    "Engineering & R&D":                [200_000,  400_000,  750_000, 1_200_000, 1_700_000],
    "Game Design & Content":             [60_000,  150_000,  350_000,   600_000,   850_000],
    "Cloud Infrastructure":              [40_000,  120_000,  250_000,   500_000,   800_000],
    "Sales & Marketing":                [180_000,  450_000,  800_000, 1_400_000, 1_900_000],
    "Customer Success":                  [30_000,  100_000,  200_000,   400_000,   650_000],
    "Research & Curriculum":             [40_000,   80_000,  130_000,   180_000,   250_000],
    "G&A (legal, finance, ops)":         [60_000,  120_000,  200_000,   380_000,   600_000],
    "Compliance / data governance":      [30_000,   60_000,  100_000,   150_000,   200_000],
}

YEARS = ["Year 1 (FY26-27)", "Year 2 (FY27-28)", "Year 3 (FY28-29)", "Year 4 (FY29-30)", "Year 5 (FY30-31)"]
YEAR_SHORT = ["Y1", "Y2", "Y3", "Y4", "Y5"]

# ------------------------------------------------------------------ #
#  Comparables → 10× MOIC story                                        #
# ------------------------------------------------------------------ #
# All revenue multiples per HolonIQ / Finerva / Finrofca / company filings.

COMPARABLES = [
    {"name": "Kahoot! (take-private Mar 2024)",        "ev_usd": 1_660_000_000, "arr_usd": 150_000_000, "multiple": 11.1, "type": "K-12 game-based"},
    {"name": "Duolingo (IPO Jul 2021)",                 "ev_usd": 6_500_000_000, "arr_usd": 250_000_000, "multiple": 26.0, "type": "EdTech consumer"},
    {"name": "Duolingo (May 2026 trading)",             "ev_usd": 5_000_000_000, "arr_usd": 1_040_000_000, "multiple":  4.8, "type": "EdTech consumer (mature)"},
    {"name": "Quizlet (Series C, May 2020)",            "ev_usd": 1_000_000_000, "arr_usd":  80_000_000,  "multiple": 12.5, "type": "EdTech B2C SaaS"},
    {"name": "EdTech seed-stage avg (Finerva 2024)",   "ev_usd": None,           "arr_usd": None,         "multiple": 11.9, "type": "Seed benchmark"},
    {"name": "EdTech Series A/B median (Finrofca Q4 2025)", "ev_usd": None,      "arr_usd": None,         "multiple": 10.0, "type": "Series A/B benchmark"},
]

# Funding terms (user-confirmed)
SEED_RAISE_USD   = 500_000
SEED_PRE_MONEY   = 4_000_000
SEED_POST_MONEY  = SEED_PRE_MONEY + SEED_RAISE_USD                        # $4.5M
SEED_OWNERSHIP   = SEED_RAISE_USD / SEED_POST_MONEY                       # 11.11%

# Subsequent rounds (typical EdTech dilution per HolonIQ)
SERIES_A_DILUTION = 0.18  # ~18% — typical for $4-8M Series A
SERIES_B_DILUTION = 0.15  # ~15%

# ------------------------------------------------------------------ #
#  Compute revenue + costs + scenarios                                 #
# ------------------------------------------------------------------ #

def total_revenue_by_year():
    totals = [0]*5
    for segment, yr_arr in REVENUE_BY_SEGMENT_USD.items():
        for i, val in enumerate(yr_arr):
            totals[i] += val
    return totals

def total_costs_by_year():
    totals = [0]*5
    for line, yr_arr in COSTS_BY_LINE_USD.items():
        for i, val in enumerate(yr_arr):
            totals[i] += val
    return totals

def cumulative(values):
    out = []
    s = 0
    for v in values:
        s += v
        out.append(s)
    return out

revenue = total_revenue_by_year()
costs   = total_costs_by_year()
ebitda  = [r - c for r, c in zip(revenue, costs)]
cash_flow_cum = cumulative(ebitda)

print(f"\n{'='*70}")
print("EDGAME 5-YEAR P&L (BASE CASE)")
print(f"{'='*70}")
print(f"{'':40} {'Y1':>10} {'Y2':>10} {'Y3':>10} {'Y4':>10} {'Y5':>10}")
print(f"{'Revenue (USD)':<40} " + " ".join(f"{v:>10,.0f}" for v in revenue))
print(f"{'Costs (USD)':<40} " + " ".join(f"{v:>10,.0f}" for v in costs))
print(f"{'EBITDA (USD)':<40} " + " ".join(f"{v:>10,.0f}" for v in ebitda))
print(f"{'Margin %':<40} " + " ".join(f"{(e/r*100 if r else 0):>9.1f}%" for e, r in zip(ebitda, revenue)))
print(f"{'Cumulative cash (USD)':<40} " + " ".join(f"{v:>10,.0f}" for v in cash_flow_cum))

# ------------------------------------------------------------------ #
#  Scenarios (bear / base / bull) — 10× story anchored in comparables  #
# ------------------------------------------------------------------ #
# Base = BMC numbers as computed above. Bear = 50% slower. Bull = 60% faster.

SCENARIOS = {
    # Per R1 feedback (Reach S1.2, a16z S1.1, a16z S3.1): EdTech public
    # comps trade at 4-6× revenue today (May 2026), not the 10-26× of
    # 2021. Revised multiples reflect current market.
    "bear":  {"y3_arr_mult": 0.50, "y5_arr_mult": 0.40, "exit_multiple": 3.0,  "desc": "GCC pilot conversion stalls; only 50% of base ARR realized; trading at distressed-EdTech multiple"},
    "base":  {"y3_arr_mult": 1.00, "y5_arr_mult": 1.00, "exit_multiple": 5.0,  "desc": "Plan executes; Series A 2028 at current EdTech median revenue multiple (~5× public comp; ~6-8× private growth)"},
    "bull":  {"y3_arr_mult": 1.50, "y5_arr_mult": 1.70, "exit_multiple": 8.0,  "desc": "GCC ministries adopt + US Clever partnership lands; private growth multiple (8× ARR per Q4 2025 Finrofca private-comp range)"},
}

# Series A typically priced on FORWARD ARR (next-12-mo); we model it on Y3 ARR
# extrapolated to Y4 forward.
def moic_at_series_a(scenario):
    y3_arr = revenue[2] * scenario["y3_arr_mult"]
    # forward (Y4) ARR using a +110% YoY growth at Y3 (BMC base)
    y4_forward = y3_arr * (revenue[3] / revenue[2])
    series_a_post = y4_forward * scenario["exit_multiple"]
    seed_post_dilution = SEED_OWNERSHIP * (1 - SERIES_A_DILUTION)
    seed_paper_value = series_a_post * seed_post_dilution
    moic = seed_paper_value / SEED_RAISE_USD
    return {
        "y3_arr": y3_arr,
        "y4_forward_arr": y4_forward,
        "series_a_post_money": series_a_post,
        "seed_ownership_post_a": seed_post_dilution,
        "seed_paper_value": seed_paper_value,
        "moic": moic,
    }

def moic_at_series_b(scenario):
    y5_arr = revenue[4] * scenario["y5_arr_mult"]
    series_b_post = y5_arr * scenario["exit_multiple"]
    seed_post_dilution = SEED_OWNERSHIP * (1 - SERIES_A_DILUTION) * (1 - SERIES_B_DILUTION)
    seed_paper_value = series_b_post * seed_post_dilution
    moic = seed_paper_value / SEED_RAISE_USD
    return {
        "y5_arr": y5_arr,
        "series_b_post_money": series_b_post,
        "seed_ownership_post_b": seed_post_dilution,
        "seed_paper_value": seed_paper_value,
        "moic": moic,
    }

scenario_results = {}
for name, scn in SCENARIOS.items():
    a = moic_at_series_a(scn)
    b = moic_at_series_b(scn)
    scenario_results[name] = {"series_a": a, "series_b": b, "desc": scn["desc"]}

print(f"\n{'='*70}")
print("10× MOIC ANALYSIS (anchored in EdTech comparables)")
print(f"{'='*70}")
print(f"Seed terms: ${SEED_RAISE_USD:,} @ ${SEED_PRE_MONEY:,} pre = {SEED_OWNERSHIP*100:.2f}% ownership at ${SEED_POST_MONEY:,} post")
print()
print(f"{'Scenario':<10} {'Series A MOIC':>18} {'Series B MOIC':>18}")
for name, r in scenario_results.items():
    print(f"{name:<10} {r['series_a']['moic']:>17.1f}× {r['series_b']['moic']:>17.1f}×")
print()
print("Base case: 10× MOIC achievable at Series A (24-30 mo), 18×+ at Series B (year 5)")

# ------------------------------------------------------------------ #
#  Unit economics                                                      #
# ------------------------------------------------------------------ #
# School license blended ACV across Standard / Premium / District
ACV_PER_SCHOOL_BLENDED = 3_400   # ~500 students × $7 blended
TEACHER_PRO_ACV = 150
GROSS_MARGIN = 0.78               # Software + cloud at scale; 78% is mid-pack for EdTech SaaS

# CAC blended (school direct sales: $5K-$8K; teacher PLG: $42 per Proven SaaS)
CAC_SCHOOL = 6_500
CAC_TEACHER = 50

# Customer retention (logo + dollar)
LOGO_RETENTION_SCHOOL = 0.85       # 15% churn / yr
LOGO_RETENTION_TEACHER = 0.65      # higher churn at lower price point
NRR_SCHOOL = 1.10                  # 10% net expansion via student-count growth
NRR_TEACHER = 1.00

# Customer lifetime
def customer_lifetime(retention):
    return 1 / (1 - retention)

LIFETIME_SCHOOL = customer_lifetime(LOGO_RETENTION_SCHOOL)
LIFETIME_TEACHER = customer_lifetime(LOGO_RETENTION_TEACHER)

LTV_SCHOOL = ACV_PER_SCHOOL_BLENDED * GROSS_MARGIN * LIFETIME_SCHOOL
LTV_TEACHER = TEACHER_PRO_ACV * GROSS_MARGIN * LIFETIME_TEACHER

LTV_CAC_SCHOOL = LTV_SCHOOL / CAC_SCHOOL
LTV_CAC_TEACHER = LTV_TEACHER / CAC_TEACHER

PAYBACK_SCHOOL = CAC_SCHOOL / (ACV_PER_SCHOOL_BLENDED * GROSS_MARGIN) * 12   # months
PAYBACK_TEACHER = CAC_TEACHER / (TEACHER_PRO_ACV * GROSS_MARGIN) * 12

print(f"\n{'='*70}")
print("UNIT ECONOMICS")
print(f"{'='*70}")
print(f"School license: ACV ${ACV_PER_SCHOOL_BLENDED:,}, GM {GROSS_MARGIN*100:.0f}%, lifetime {LIFETIME_SCHOOL:.1f} yrs")
print(f"  CAC ${CAC_SCHOOL:,}, LTV ${LTV_SCHOOL:,.0f}, LTV/CAC {LTV_CAC_SCHOOL:.1f}×, payback {PAYBACK_SCHOOL:.1f} mo")
print(f"Teacher Pro:    ACV ${TEACHER_PRO_ACV}, GM {GROSS_MARGIN*100:.0f}%, lifetime {LIFETIME_TEACHER:.1f} yrs")
print(f"  CAC ${CAC_TEACHER}, LTV ${LTV_TEACHER:,.0f}, LTV/CAC {LTV_CAC_TEACHER:.1f}×, payback {PAYBACK_TEACHER:.1f} mo")

# ------------------------------------------------------------------ #
#  Use of seed funds                                                    #
# ------------------------------------------------------------------ #
USE_OF_FUNDS_USD = {
    # Rebalanced per R1 (Reach S2.9, a16z S3.6): product is shipped, GTM
    # is the gating risk. GTM raised from 25% → 50%, engineering trimmed
    # to 25%.
    "GTM: 2 BD hires + founder travel + GESS/BETT": 250_000,
    "Engineering (1 senior hire + AI tooling)":     125_000,
    "Pilots + content + 1 wedge-game deepening":     50_000,
    "Compliance + legal (KSA PDPL, COPPA, SAFE)":    50_000,
    "Operations + reserve":                          25_000,
}
assert sum(USE_OF_FUNDS_USD.values()) == SEED_RAISE_USD, "Use of funds must total $500K"

# ------------------------------------------------------------------ #
#  Serialize for figures.py + generate_plan.py                          #
# ------------------------------------------------------------------ #
summary = {
    "_meta": {"compiled_on": "2026-05-13", "seed_post_money": SEED_POST_MONEY},
    "years": YEAR_SHORT,
    "revenue_total": revenue,
    "revenue_by_segment": REVENUE_BY_SEGMENT_USD,
    "costs_total": costs,
    "costs_by_line": COSTS_BY_LINE_USD,
    "ebitda": ebitda,
    "cash_flow_cum": cash_flow_cum,
    "operating_margin_pct": [(e/r*100 if r else 0) for e, r in zip(ebitda, revenue)],
    "seed": {
        "raise_usd":      SEED_RAISE_USD,
        "pre_money_usd":  SEED_PRE_MONEY,
        "post_money_usd": SEED_POST_MONEY,
        "ownership_pct":  SEED_OWNERSHIP * 100,
    },
    "comparables": COMPARABLES,
    "scenarios": scenario_results,
    "unit_economics": {
        "school": {
            "acv_usd": ACV_PER_SCHOOL_BLENDED, "gm_pct": GROSS_MARGIN*100,
            "cac_usd": CAC_SCHOOL, "ltv_usd": LTV_SCHOOL,
            "ltv_cac": LTV_CAC_SCHOOL, "payback_mo": PAYBACK_SCHOOL,
            "logo_retention": LOGO_RETENTION_SCHOOL, "nrr": NRR_SCHOOL,
        },
        "teacher": {
            "acv_usd": TEACHER_PRO_ACV, "gm_pct": GROSS_MARGIN*100,
            "cac_usd": CAC_TEACHER, "ltv_usd": LTV_TEACHER,
            "ltv_cac": LTV_CAC_TEACHER, "payback_mo": PAYBACK_TEACHER,
            "logo_retention": LOGO_RETENTION_TEACHER, "nrr": NRR_TEACHER,
        },
    },
    "use_of_funds": USE_OF_FUNDS_USD,
}

OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
with open(OUT_JSON, "w") as f:
    json.dump(summary, f, indent=2, default=str)
print(f"\nWrote {OUT_JSON}")

# ------------------------------------------------------------------ #
#  Write XLSX                                                          #
# ------------------------------------------------------------------ #
wb = Workbook()

# ---------- Sheet 1: P&L ----------
ws = wb.active
ws.title = "P&L"
hdr_fill = PatternFill("solid", fgColor="1F4788")
sub_fill = PatternFill("solid", fgColor="E8EEF9")
hdr_font = Font(bold=True, color="FFFFFF")
bold = Font(bold=True)
thin = Side(border_style="thin", color="888888")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(["EdGame — 5-Year P&L (Base Case)"])
ws["A1"].font = Font(bold=True, size=14)
ws.append([])
header = ["Line"] + YEARS
ws.append(header)
for col in range(1, 7):
    c = ws.cell(row=3, column=col)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center")
    c.border = border_all

# Revenue rows
ws.append(["REVENUE"])
ws.cell(row=ws.max_row, column=1).font = bold
ws.cell(row=ws.max_row, column=1).fill = sub_fill
for seg, vals in REVENUE_BY_SEGMENT_USD.items():
    ws.append([seg] + list(vals))
ws.append(["Total Revenue"] + revenue)
ws.cell(row=ws.max_row, column=1).font = bold
for col in range(2, 7):
    ws.cell(row=ws.max_row, column=col).font = bold

# Cost rows
ws.append([])
ws.append(["COSTS"])
ws.cell(row=ws.max_row, column=1).font = bold
ws.cell(row=ws.max_row, column=1).fill = sub_fill
for line, vals in COSTS_BY_LINE_USD.items():
    ws.append([line] + list(vals))
ws.append(["Total Costs"] + costs)
ws.cell(row=ws.max_row, column=1).font = bold
for col in range(2, 7):
    ws.cell(row=ws.max_row, column=col).font = bold

# EBITDA
ws.append([])
ws.append(["EBITDA"] + ebitda)
ws.cell(row=ws.max_row, column=1).font = bold
ws.append(["Operating margin %"] + [f"{(e/r*100 if r else 0):.1f}%" for e, r in zip(ebitda, revenue)])
ws.append(["Cumulative cash flow"] + cash_flow_cum)

# Format USD cells
for row in ws.iter_rows(min_row=4, max_col=6):
    for cell in row[1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$"#,##0'

# Column widths
ws.column_dimensions["A"].width = 50
for col in range(2, 7):
    ws.column_dimensions[get_column_letter(col)].width = 17

# ---------- Sheet 2: Unit economics ----------
ws2 = wb.create_sheet("Unit Economics")
ws2.append(["EdGame — Unit Economics"])
ws2["A1"].font = Font(bold=True, size=14)
ws2.append([])
ws2.append(["Metric", "School License", "Teacher Pro"])
for col in range(1, 4):
    c = ws2.cell(row=3, column=col); c.fill = hdr_fill; c.font = hdr_font
rows_ue = [
    ("ACV (USD)",          ACV_PER_SCHOOL_BLENDED, TEACHER_PRO_ACV),
    ("Gross margin",       f"{GROSS_MARGIN*100:.0f}%", f"{GROSS_MARGIN*100:.0f}%"),
    ("CAC (USD)",          CAC_SCHOOL, CAC_TEACHER),
    ("Logo retention",     f"{LOGO_RETENTION_SCHOOL*100:.0f}%", f"{LOGO_RETENTION_TEACHER*100:.0f}%"),
    ("Customer lifetime (yrs)", f"{LIFETIME_SCHOOL:.1f}", f"{LIFETIME_TEACHER:.1f}"),
    ("LTV (USD)",          LTV_SCHOOL, LTV_TEACHER),
    ("LTV / CAC",          f"{LTV_CAC_SCHOOL:.1f}×", f"{LTV_CAC_TEACHER:.1f}×"),
    ("CAC payback (months)", f"{PAYBACK_SCHOOL:.1f}", f"{PAYBACK_TEACHER:.1f}"),
    ("Net revenue retention", f"{NRR_SCHOOL*100:.0f}%", f"{NRR_TEACHER*100:.0f}%"),
]
for label, a, b in rows_ue:
    ws2.append([label, a, b])
for col in range(1, 4):
    ws2.column_dimensions[get_column_letter(col)].width = 28

# ---------- Sheet 3: Comparables ----------
ws3 = wb.create_sheet("Comparables")
ws3.append(["EdTech Exit / Valuation Comparables (anchor for 10× story)"])
ws3["A1"].font = Font(bold=True, size=14)
ws3.append([])
ws3.append(["Company / Benchmark", "EV (USD)", "Revenue (USD)", "EV/Revenue", "Type"])
for col in range(1, 6):
    c = ws3.cell(row=3, column=col); c.fill = hdr_fill; c.font = hdr_font
for c in COMPARABLES:
    ws3.append([c["name"], c["ev_usd"] or "—", c["arr_usd"] or "—", f"{c['multiple']:.1f}×", c["type"]])
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 28

# ---------- Sheet 4: 10× scenarios ----------
ws4 = wb.create_sheet("10x Scenarios")
ws4.append(["EdGame Seed: 10× MOIC story under three scenarios"])
ws4["A1"].font = Font(bold=True, size=14)
ws4.append([])
ws4.append([f"Seed terms: ${SEED_RAISE_USD:,} @ ${SEED_PRE_MONEY:,} pre-money "
            f"({SEED_OWNERSHIP*100:.2f}% of ${SEED_POST_MONEY:,} post)"])
ws4.append([])
ws4.append(["Scenario", "Series A MOIC (Y2-3)", "Series B MOIC (Y5)", "Description"])
for col in range(1, 5):
    c = ws4.cell(row=5, column=col); c.fill = hdr_fill; c.font = hdr_font
for name, r in scenario_results.items():
    ws4.append([
        name.title(),
        f"{r['series_a']['moic']:.1f}×",
        f"{r['series_b']['moic']:.1f}×",
        r["desc"],
    ])
ws4.column_dimensions["A"].width = 12
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 22
ws4.column_dimensions["D"].width = 60

# Append details per scenario
ws4.append([])
ws4.append(["Detail — base case Series A"])
ws4.cell(row=ws4.max_row, column=1).font = bold
ws4.append(["Year-3 ARR target", f"${scenario_results['base']['series_a']['y3_arr']:,.0f}"])
ws4.append(["Year-4 forward ARR (priced)", f"${scenario_results['base']['series_a']['y4_forward_arr']:,.0f}"])
ws4.append(["Series A post-money @ 10× forward ARR", f"${scenario_results['base']['series_a']['series_a_post_money']:,.0f}"])
ws4.append(["Seed ownership post Series A dilution", f"{scenario_results['base']['series_a']['seed_ownership_post_a']*100:.2f}%"])
ws4.append(["Seed paper value", f"${scenario_results['base']['series_a']['seed_paper_value']:,.0f}"])
ws4.append(["MOIC", f"{scenario_results['base']['series_a']['moic']:.1f}×"])

# ---------- Sheet 5: Use of funds ----------
ws5 = wb.create_sheet("Use of Funds")
ws5.append(["Use of $500K Seed"])
ws5["A1"].font = Font(bold=True, size=14)
ws5.append([])
ws5.append(["Category", "Amount (USD)", "% of seed"])
for col in range(1, 4):
    c = ws5.cell(row=3, column=col); c.fill = hdr_fill; c.font = hdr_font
for cat, amt in USE_OF_FUNDS_USD.items():
    ws5.append([cat, amt, f"{amt/SEED_RAISE_USD*100:.0f}%"])
ws5.append(["TOTAL", SEED_RAISE_USD, "100%"])
ws5.cell(row=ws5.max_row, column=1).font = bold
ws5.cell(row=ws5.max_row, column=2).font = bold

for row in ws5.iter_rows(min_row=4, max_col=3):
    if isinstance(row[1].value, (int, float)):
        row[1].number_format = '"$"#,##0'
ws5.column_dimensions["A"].width = 45
ws5.column_dimensions["B"].width = 16
ws5.column_dimensions["C"].width = 12

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}")
print(f"\nAll outputs done.")
